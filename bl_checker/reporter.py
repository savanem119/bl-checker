"""
reporter.py -- Export du rapport de controle en Excel (openpyxl) et PDF (fpdf2).
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from comparator import ControlLine, Status

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

COL_HEADERS = [
    "Rubrique",
    "Correction demandee",
    "Valeur trouvee",
    "Statut",
    "Criticite",
    "Confiance",
    "Commentaire",
]

STATUS_LABELS = {
    Status.CONFORME:     "Conforme",
    Status.AVEC_RESERVE: "Avec reserve",
    Status.PARTIEL:      "Partiel",
    Status.NON_CONFORME: "Non conforme",
    Status.A_VERIFIER:   "A verifier",
}

STATUS_COLORS_XL = {
    Status.CONFORME:     "C6EFCE",
    Status.AVEC_RESERVE: "FFEB9C",
    Status.PARTIEL:      "FFC7CE",
    Status.NON_CONFORME: "FF4C4C",
    Status.A_VERIFIER:   "BDD7EE",
}

CRIT_COLORS_XL = {
    "critique":  "FF0000",
    "important": "FF9900",
    "mineur":    "AAAAAA",
}

# ---------------------------------------------------------------------------
# Utilitaire
# ---------------------------------------------------------------------------

def _safe(s: str) -> str:
    """Remplace les caracteres hors latin-1 par des equivalents ASCII."""
    if not s:
        return ""
    # Remplacements symboliques courants
    for char, repl in [
        ("\u2014", "--"), ("\u2013", "-"),
        ("\u2019", "'"), ("\u2018", "'"),
        ("\u201c", '"'), ("\u201d", '"'),
        ("\u26a0", "(!)"),  # warning sign
        ("\u2260", "!="),   # not equal
        ("\u2264", "<="),   # less or equal
        ("\u2265", ">="),   # greater or equal
        ("\u2192", "->"),   # right arrow
        ("\u2190", "<-"),   # left arrow
        ("\u00e9", "e"), ("\u00e8", "e"), ("\u00ea", "e"), ("\u00eb", "e"),
        ("\u00e0", "a"), ("\u00e2", "a"), ("\u00e4", "a"),
        ("\u00f4", "o"), ("\u00f6", "o"),
        ("\u00f9", "u"), ("\u00fb", "u"), ("\u00fc", "u"),
        ("\u00ee", "i"), ("\u00ef", "i"),
        ("\u00e7", "c"),
        ("\u00e6", "ae"), ("\u0153", "oe"),
        ("\u00c9", "E"), ("\u00c0", "A"), ("\u00c7", "C"),
        ("\u00b0", " deg"), ("\u2026", "..."),
        ("\u00a0", " "),
    ]:
        s = s.replace(char, repl)
    return s.encode("latin-1", errors="replace").decode("latin-1")

# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

def export_excel(report: Dict[str, Any], output_path: str) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError("openpyxl requis : pip install openpyxl") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport de controle"

    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def bfont(size=10, color="000000"):
        return Font(bold=True, size=size, color=color)

    meta = report["metadata"]
    ws["A1"] = "RAPPORT DE CONTROLE BL"
    ws["A1"].font = bfont(14)
    ws.merge_cells("A1:G1")
    ws["A2"] = f"BL corrige : {meta.get('fichier_corrige', '')}"
    ws["D2"] = f"Nouveau BL : {meta.get('fichier_nouveau', '')}"
    ws["A3"] = f"Date/Heure : {meta.get('date_heure', '')}"
    ws["D3"] = f"Operateur  : {meta.get('utilisateur', '')}"
    ws["A4"] = f"Statut global : {report.get('global_status', '')}"
    ws["A4"].font = bfont(11, "CC0000")
    ws.merge_cells("A4:G4")
    ws["A5"] = " "

    header_row = 6
    for ci, h in enumerate(COL_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = bfont(10, "FFFFFF")
        cell.fill = fill("1F3864")
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        cell.border = border
    ws.row_dimensions[header_row].height = 22

    all_lines = report.get("control_lines", []) + report.get("consistency_lines", [])
    for row_offset, line in enumerate(all_lines):
        r = header_row + 1 + row_offset
        status_label = STATUS_LABELS.get(line.statut, str(line.statut))
        conf_str = f"{line.confiance:.0%}" if line.confiance is not None else ""
        row_vals = [
            line.rubrique,
            (line.correction_demandee or "")[:120],
            (line.valeur_trouvee or "")[:120],
            status_label,
            line.criticite or "",
            conf_str,
            (line.commentaire or "")[:200],
        ]
        sfill = STATUS_COLORS_XL.get(line.statut, "FFFFFF")
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            if ci == 4:
                cell.fill = fill(sfill)
                if line.statut == Status.NON_CONFORME:
                    cell.font = Font(bold=True, color="FFFFFF")
            elif ci == 5 and line.criticite == "critique":
                cell.font = Font(bold=True, color=CRIT_COLORS_XL["critique"])
            if r % 2 == 0 and ci not in (4, 5):
                cell.fill = fill("F2F2F2")
        ws.row_dimensions[r].height = 30

    col_widths = [28, 32, 32, 16, 12, 10, 40]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws["A7"]

    ws_stats = wb.create_sheet("Statistiques")
    stats = report.get("stats", {})
    ws_stats["A1"] = "STATISTIQUES"
    ws_stats["A1"].font = bfont(12)
    stat_rows = [
        ("Total points de controle", stats.get("total", 0)),
        ("Conformes",                stats.get("conforme", 0)),
        ("Avec reserve",             stats.get("avec_reserve", 0)),
        ("Partiels",                 stats.get("partiel", 0)),
        ("Non conformes",            stats.get("non_conforme", 0)),
        ("  dont Critiques NC",      stats.get("critique_nc", 0)),
        ("  dont Importants NC",     stats.get("important_nc", 0)),
        ("  dont Mineurs NC",        stats.get("mineur_nc", 0)),
        ("A verifier",               stats.get("a_verifier", 0)),
    ]
    for i, (label, val) in enumerate(stat_rows, start=2):
        ws_stats.cell(row=i, column=1, value=label)
        ws_stats.cell(row=i, column=2, value=val)
    ws_stats.column_dimensions["A"].width = 30
    ws_stats.column_dimensions["B"].width = 12
    ws_stats["A12"] = "Conclusion"
    ws_stats["A12"].font = bfont()
    ws_stats["A13"] = report.get("conclusion", "")
    ws_stats.column_dimensions["A"].width = 80

    bl_fields = report.get("bl_fields")
    if bl_fields and bl_fields.containers:
        ws_cont = wb.create_sheet("Conteneurs")
        cont_headers = ["Numero", "Type", "Sacs", "Poids brut", "Seal", "Page"]
        for ci, h in enumerate(cont_headers, start=1):
            cell = ws_cont.cell(row=1, column=ci, value=h)
            cell.font = bfont(10, "FFFFFF")
            cell.fill = fill("1F3864")
            cell.border = border
        for ri, cont in enumerate(bl_fields.containers, start=2):
            for ci, v in enumerate(
                [cont.number, cont.container_type, cont.bags,
                 cont.gross_weight, cont.seal, cont.page_num + 1], start=1
            ):
                ws_cont.cell(row=ri, column=ci, value=v).border = border
        for ci, w in enumerate([16, 10, 8, 12, 14, 6], start=1):
            ws_cont.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Export PDF
# ---------------------------------------------------------------------------

def export_pdf(report: Dict[str, Any], output_path: str) -> str:
    try:
        from fpdf import FPDF
    except ImportError as e:
        raise ImportError("fpdf2 requis : pip install fpdf2") from e

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.set_left_margin(10)
    pdf.set_right_margin(10)
    pdf.add_page()

    meta = report["metadata"]

    # En-tete
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "RAPPORT DE CONTROLE BL", ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, _safe(
        f"BL corrige : {meta.get('fichier_corrige','')}  |  Nouveau BL : {meta.get('fichier_nouveau','')}"
    ), ln=True)
    pdf.cell(0, 5, _safe(
        f"Date/Heure : {meta.get('date_heure','')}  |  Operateur : {meta.get('utilisateur','')}"
    ), ln=True)
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(180, 0, 0)
    pdf.cell(0, 7, _safe(f"Statut global : {report.get('global_status', '')}"), ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    stats = report.get("stats", {})
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, _safe(
        f"Total : {stats.get('total',0)} points"
        f" | Conformes : {stats.get('conforme',0)}"
        f" | Non conformes : {stats.get('non_conforme',0)}"
        f" | A verifier : {stats.get('a_verifier',0)}"
    ), ln=True)
    pdf.ln(4)

    STATUS_SHORT = {
        Status.CONFORME:     "OK",
        Status.AVEC_RESERVE: "Reserve",
        Status.PARTIEL:      "Partiel",
        Status.NON_CONFORME: "NON CONF.",
        Status.A_VERIFIER:   "A VERIF.",
    }
    STATUS_RGB = {
        Status.CONFORME:     (198, 239, 206),
        Status.AVEC_RESERVE: (255, 235, 156),
        Status.PARTIEL:      (255, 199, 206),
        Status.NON_CONFORME: (255, 100, 100),
        Status.A_VERIFIER:   (189, 215, 238),
    }

    col_w = [48, 35, 35, 22, 18, 30]
    hdr_names = ["Rubrique", "Demande", "Trouve", "Statut", "Crit.", "Commentaire"]

    def draw_table_header():
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(31, 56, 100)
        pdf.set_text_color(255, 255, 255)
        for i, h in enumerate(hdr_names):
            pdf.cell(col_w[i], 6, h, border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

    draw_table_header()

    all_lines = report.get("control_lines", []) + report.get("consistency_lines", [])

    for idx, line in enumerate(all_lines):
        rgb = STATUS_RGB.get(line.statut, (255, 255, 255))
        pdf.set_fill_color(*rgb)
        pdf.set_font("Helvetica", "", 7)

        row_data = [
            _safe((line.rubrique or "")[:45]),
            _safe((line.correction_demandee or "")[:35]),
            _safe((line.valeur_trouvee or "")[:35]),
            _safe(STATUS_SHORT.get(line.statut, str(line.statut))),
            _safe((line.criticite or "")[:10]),
            _safe((line.commentaire or "")[:50]),
        ]
        row_h = 6

        if pdf.get_y() + row_h > pdf.page_break_trigger:
            pdf.add_page()
            draw_table_header()
            pdf.set_font("Helvetica", "", 7)
            pdf.set_fill_color(*rgb)

        do_fill = line.statut != Status.CONFORME or idx % 2 == 0
        for val, cw in zip(row_data, col_w):
            pdf.cell(cw, row_h, val, border=1, fill=do_fill)
        pdf.ln()

    # Conclusion
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 8, "CONCLUSION OPERATIONNELLE", ln=True)
    pdf.set_font("Helvetica", "", 9)
    for line_txt in report.get("conclusion", "").split("\n"):
        pdf.multi_cell(pdf.epw, 5, _safe(line_txt))
    pdf.ln(3)

    # Conteneurs
    bl_fields = report.get("bl_fields")
    if bl_fields and bl_fields.containers:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 8, "DETAIL CONTENEURS (nouveau BL)", ln=True)
        cw2 = [32, 16, 16, 22, 28]
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(31, 56, 100)
        pdf.set_text_color(255, 255, 255)
        for h, cw in zip(["Numero", "Type", "Sacs", "Poids brut", "Seal"], cw2):
            pdf.cell(cw, 6, h, border=1, fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 8)
        for ci, cont in enumerate(bl_fields.containers):
            if ci % 2 == 0:
                pdf.set_fill_color(255, 255, 255)
            else:
                pdf.set_fill_color(242, 242, 242)
            vals2 = [
                _safe(cont.number),
                _safe(cont.container_type or ""),
                _safe(str(cont.bags or "")),
                _safe(str(cont.gross_weight or "")),
                _safe(cont.seal or ""),
            ]
            for val, cw in zip(vals2, cw2):
                pdf.cell(cw, 5, val, border=1, fill=True)
            pdf.ln()

    pdf.output(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Point d'entree generique
# ---------------------------------------------------------------------------

def export_report(
    report: Dict[str, Any],
    output_dir: str = ".",
    prefix: str = "rapport_bl",
    formats: List[str] = None,
) -> Dict[str, str]:
    if formats is None:
        formats = ["xlsx", "pdf"]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.join(output_dir, f"{prefix}_{ts}")
    result = {}
    if "xlsx" in formats:
        path_xl = base + ".xlsx"
        export_excel(report, path_xl)
        result["xlsx"] = path_xl
    if "pdf" in formats:
        path_pdf = base + ".pdf"
        export_pdf(report, path_pdf)
        result["pdf"] = path_pdf
    return result
